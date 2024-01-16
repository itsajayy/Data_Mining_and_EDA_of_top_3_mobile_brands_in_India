from bs4 import BeautifulSoup
import openpyxl
import os

# Read the HTML file
#Save the amazon website link by clicking CTRL+S, into the same directory/file location as your Code.
# #Add your HTML file that you saved in the Quotes
with open(" ", "r", encoding="utf-8") as file: 


    html_content = file.read()

soup = BeautifulSoup(html_content, 'html.parser')

# Find all div elements with the specified class
product_divs = soup.find_all("div", class_="puis-card-container s-card-container s-overflow-hidden aok-relative puis-include-content-margin puis puis-v2q9dos4w4qqgu20zj8pkw7yd24 s-latency-cf-section puis-card-border")

# Initialize lists to store data
product_names = []
product_costs = []
product_ratings = []
product_reviews = []

# Loop through the found div elements
for div in product_divs:
    # Find the product name
    product_name = div.find("span", class_="a-size-medium a-color-base a-text-normal")
    product_name = product_name.text if product_name else "NA"
    product_names.append(product_name)
    
    # Find the product cost
    product_cost = div.find("span", class_="a-price-whole")
    product_cost = product_cost.text if product_cost else "NA"
    product_costs.append(product_cost)
    
    # Find the product ratings
    product_rating = div.find("span", class_="a-icon-alt")
    product_rating = product_rating.text if product_rating else "NA"
    product_ratings.append(product_rating)
    
    # Find the product review
    product_review = div.find("span", class_="a-size-base s-underline-text")
    product_review = product_review.text if product_review else "NA"
    product_reviews.append(product_review)

# Open the existing Excel file or create a new one if it doesn't exist
excel_file = "samsung_phone_data.xlsx"
workbook = openpyxl.load_workbook(excel_file) if os.path.exists(excel_file) else openpyxl.Workbook()
sheet = workbook.active

# If the Excel file is empty, write headers
if sheet.max_row == 1:
    sheet.append(["Product Name", "Product Cost", "Product Ratings", "Product Review"])

# Write Data
for i in range(len(product_names)):
    row_data = [product_names[i], product_costs[i], product_ratings[i], product_reviews[i]]
    sheet.append(row_data)

# Save the Excel file
workbook.save(excel_file)

print(f"Data has been appended to {excel_file}")
